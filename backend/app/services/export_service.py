"""Export service — CSV / Excel / PDF generation."""

import calendar
import csv
import io
from datetime import date

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.day_program import DayProgram
from app.models.master import TimeBlockMaster
from app.models.schedule import Schedule, ScheduleAssignment
from app.models.staff import Staff
from app.services.schedule_service import TIME_BLOCK_ORDER

WEEKDAYS_JP = ["月", "火", "水", "木", "金", "土", "日"]


async def _load_export_data(db: AsyncSession, schedule: Schedule) -> dict:
    """Load all data needed for export (shared by CSV/Excel/PDF)."""
    year, month = map(int, schedule.year_month.split("-"))
    _, last_day = calendar.monthrange(year, month)

    # Fetch staff
    staff_result = await db.execute(
        select(Staff).where(Staff.is_active == True).order_by(Staff.name)  # noqa: E712
    )
    staff_list = staff_result.scalars().all()

    # Fetch time blocks
    tb_result = await db.execute(select(TimeBlockMaster).order_by(TimeBlockMaster.sort_order))
    time_blocks = tb_result.scalars().all()
    tb_display = {tb.code: tb.display_name for tb in time_blocks}

    # Fetch assignments
    assign_result = await db.execute(
        select(ScheduleAssignment).where(ScheduleAssignment.schedule_id == schedule.id)
    )
    assignments = assign_result.scalars().all()
    assign_index: dict[tuple, ScheduleAssignment] = {}
    for a in assignments:
        assign_index[(a.date, a.time_block, str(a.staff_id))] = a

    # Fetch day programs
    dp_result = await db.execute(
        select(DayProgram).where(DayProgram.schedule_id == schedule.id)
    )
    day_programs = dp_result.scalars().all()
    dp_index: dict[tuple, DayProgram] = {}
    for dp in day_programs:
        dp_index[(dp.date, dp.time_block)] = dp

    return {
        "year": year,
        "month": month,
        "last_day": last_day,
        "staff_list": staff_list,
        "tb_display": tb_display,
        "assign_index": assign_index,
        "dp_index": dp_index,
    }


def _get_cell_text(assign_index: dict, current_date: date, block_code: str, staff_id: str) -> str:
    a = assign_index.get((current_date, block_code, staff_id))
    if not a:
        return ""
    parts = []
    if a.task_type_code:
        parts.append(a.task_type_code)
    if a.display_text:
        parts.append(a.display_text)
    return " ".join(parts)


async def generate_csv(db: AsyncSession, schedule: Schedule) -> str:
    data = await _load_export_data(db, schedule)

    output = io.StringIO()
    writer = csv.writer(output)

    header = ["日付", "曜日", "時間帯", "DNC", "予定"]
    header.extend([s.name for s in data["staff_list"]])
    writer.writerow(header)

    for day_num in range(1, data["last_day"] + 1):
        current_date = date(data["year"], data["month"], day_num)
        weekday = WEEKDAYS_JP[current_date.weekday()]
        for block_code in TIME_BLOCK_ORDER:
            dp = data["dp_index"].get((current_date, block_code))
            row = [
                f"{data['month']}/{day_num}",
                weekday,
                data["tb_display"].get(block_code, block_code),
                dp.program_title or "" if dp else "",
                dp.summary_text or "" if dp else "",
            ]
            for staff in data["staff_list"]:
                row.append(_get_cell_text(data["assign_index"], current_date, block_code, str(staff.id)))
            writer.writerow(row)

    return output.getvalue()


async def generate_excel(db: AsyncSession, schedule: Schedule) -> bytes:
    """Generate an Excel (.xlsx) file for the schedule."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    data = await _load_export_data(db, schedule)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data['year']}年{data['month']}月"

    # Styles
    header_font = Font(bold=True, size=9)
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    weekend_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(data["staff_list"]))
    title_cell = ws.cell(row=1, column=1, value=f"シフト表 {data['year']}年{data['month']}月")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Header row
    row_idx = 3
    headers = ["日付", "曜日", "時間帯", "DNC", "予定"]
    headers.extend([s.name for s in data["staff_list"]])

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    for i in range(len(data["staff_list"])):
        ws.column_dimensions[get_column_letter(6 + i)].width = 12

    # Data rows
    row_idx = 4
    for day_num in range(1, data["last_day"] + 1):
        current_date = date(data["year"], data["month"], day_num)
        weekday = WEEKDAYS_JP[current_date.weekday()]
        is_weekend = current_date.weekday() >= 5

        for bi, block_code in enumerate(TIME_BLOCK_ORDER):
            dp = data["dp_index"].get((current_date, block_code))
            row_data = [
                f"{data['month']}/{day_num}" if bi == 0 else "",
                weekday if bi == 0 else "",
                data["tb_display"].get(block_code, block_code),
                dp.program_title or "" if dp else "",
                dp.summary_text or "" if dp else "",
            ]
            for staff in data["staff_list"]:
                row_data.append(_get_cell_text(data["assign_index"], current_date, block_code, str(staff.id)))

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if is_weekend:
                    cell.fill = weekend_fill

            row_idx += 1

    # Freeze panes
    ws.freeze_panes = "F4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_pdf(db: AsyncSession, schedule: Schedule) -> bytes:
    """Generate a PDF file for the schedule."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    data = await _load_export_data(db, schedule)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(
        f"<b>シフト表 {data['year']}年{data['month']}月</b>",
        styles["Title"],
    ))
    elements.append(Spacer(1, 5 * mm))

    # Build table data (split by week for readability)
    staff_names = [s.name for s in data["staff_list"]]
    max_staff_cols = min(len(staff_names), 10)  # Limit columns for PDF

    header = ["日", "曜", "時間帯", "DNC"]
    header.extend(staff_names[:max_staff_cols])
    if len(staff_names) > max_staff_cols:
        header.append(f"他{len(staff_names) - max_staff_cols}名")

    table_data = [header]

    for day_num in range(1, data["last_day"] + 1):
        current_date = date(data["year"], data["month"], day_num)
        weekday = WEEKDAYS_JP[current_date.weekday()]

        for bi, block_code in enumerate(TIME_BLOCK_ORDER):
            dp = data["dp_index"].get((current_date, block_code))
            row = [
                f"{day_num}" if bi == 0 else "",
                weekday if bi == 0 else "",
                data["tb_display"].get(block_code, block_code),
                (dp.program_title or "") if dp else "",
            ]
            for staff in data["staff_list"][:max_staff_cols]:
                row.append(_get_cell_text(data["assign_index"], current_date, block_code, str(staff.id)))
            if len(staff_names) > max_staff_cols:
                row.append("")
            table_data.append(row)

    # Column widths
    num_cols = len(header)
    avail_width = landscape(A4)[0] - 20 * mm
    fixed_widths = [18 * mm, 12 * mm, 25 * mm, 25 * mm]  # day, weekday, time, DNC
    remaining = avail_width - sum(fixed_widths)
    staff_col_width = remaining / max(num_cols - 4, 1)
    col_widths = fixed_widths + [staff_col_width] * (num_cols - 4)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Style
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTSIZE", (0, 0), (-1, -1), 6),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]

    # Highlight weekends
    row_idx = 1
    for day_num in range(1, data["last_day"] + 1):
        current_date = date(data["year"], data["month"], day_num)
        if current_date.weekday() >= 5:
            for _ in TIME_BLOCK_ORDER:
                style_cmds.append(
                    ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#FEF3C7"))
                )
                row_idx += 1
        else:
            row_idx += len(TIME_BLOCK_ORDER)

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    # Legend
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph(
        f"<i>出力日時: {date.today().isoformat()} | 職員数: {len(staff_names)} | "
        f"ステータス: {schedule.status}</i>",
        styles["Normal"],
    ))

    doc.build(elements)
    return buf.getvalue()
